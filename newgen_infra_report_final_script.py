import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Configuration
GRAFANA_URL = "https://monitoring.bhartiaxa.com"
API_KEY = "#############"
DATASOURCE_UID = "0wSD8eb7z"
NAMESPACE = "default"
OUTPUT_PATH = "/mnt/c/Users/Parikshit Kudalkar/Downloads"
CONTAINERS = [
    "brmsinstanceweb", "brmsinstanceejb", "od110services",
    "ibps5sp3uiweb", "ibps5sp3uiejb", "ibps5sp3aiweb", "ibps5sp3aiejb"
]

def format_percentage(value):
    """Format percentage values to XX.XX% format"""
    if isinstance(value, (int, float)):
        return f"{value:.2f}%"
    return value

def ensure_output_directory():
    """Ensure the output directory exists"""
    if not os.path.exists(OUTPUT_PATH):
        os.makedirs(OUTPUT_PATH)
        print(f"Created output directory: {OUTPUT_PATH}")

def fetch_pod_names(container_name):
    """Fetch pod names for the given container with retry logic"""
    max_retries = 3
    retry_delay = 1
    query = f'kube_pod_container_info{{container="{container_name}", namespace="{NAMESPACE}"}}'
    
    for attempt in range(max_retries):
        try:
            response = requests.get(
                f"{GRAFANA_URL}/api/datasources/proxy/uid/{DATASOURCE_UID}/api/v1/query",
                headers={"Authorization": f"Bearer {API_KEY}"},
                params={"query": query},
                timeout=10
            )
            response.raise_for_status()
            pods = [result['metric']['pod'] for result in response.json().get('data', {}).get('result', [])]
            return f"({'|'.join(pods)})" if pods else ""
        except (requests.exceptions.RequestException, KeyError) as e:
            if attempt == max_retries - 1:
                print(f"Failed to fetch pods for {container_name}: {str(e)}")
                return ""
            time.sleep(retry_delay)
    return ""

def fetch_metrics(container_name, pod_regex, start_time, end_time):
    """Fetch metrics with improved accuracy and error handling"""
    try:
        # CPU Query - returns values as decimals (e.g., 0.9479 for 94.79%)
        cpu_query = f'''
        sum(rate(container_cpu_usage_seconds_total{{
            namespace="{NAMESPACE}",
            pod=~"{pod_regex}",
            container="{container_name}",
            image!=""
        }}[5m])) by (container)
        /
        sum(kube_pod_container_resource_limits{{
            namespace="{NAMESPACE}",
            pod=~"{pod_regex}",
            container="{container_name}",
            resource="cpu"
        }}) by (container)
        '''

        # Memory Query - returns values as decimals
        mem_query = f'''
        sum(container_memory_working_set_bytes{{
            namespace="{NAMESPACE}",
            pod=~"{pod_regex}",
            container="{container_name}",
            image!=""
        }}) by (container)
        /
        sum(kube_pod_container_resource_limits{{
            namespace="{NAMESPACE}",
            pod=~"{pod_regex}",
            container="{container_name}",
            resource="memory"
        }}) by (container)
        '''

        params = {
            "start": start_time.timestamp(),
            "end": end_time.timestamp(),
            "step": "1m"
        }

        def fetch_data(query):
            for attempt in range(3):
                try:
                    response = requests.get(
                        f"{GRAFANA_URL}/api/datasources/proxy/uid/{DATASOURCE_UID}/api/v1/query_range",
                        headers={"Authorization": f"Bearer {API_KEY}"},
                        params={**params, "query": query},
                        timeout=15
                    )
                    response.raise_for_status()
                    return response.json()
                except requests.exceptions.RequestException:
                    if attempt == 2:
                        raise
                    time.sleep(1)

        cpu_data = fetch_data(cpu_query)
        mem_data = fetch_data(mem_query)

        def process_values(data):
            values = []
            for result in data.get('data', {}).get('result', []):
                for point in result.get('values', []):
                    try:
                        val = float(point[1]) * 100  # Convert to percentage
                        if val >= 0:
                            values.append(val)
                    except (ValueError, TypeError):
                        continue
            return values

        cpu_values = process_values(cpu_data)
        mem_values = process_values(mem_data)

        return cpu_values, mem_values

    except Exception as e:
        print(f"Error fetching metrics for {container_name}: {str(e)}")
        return None, None

def generate_report():
    """Generate comprehensive report with accurate metrics"""
    end_time = datetime.now()
    start_time = end_time - timedelta(hours=6)
    report_date = end_time.strftime("%d-%m-%y")
    
    report = []
    for container in CONTAINERS:
        pod_regex = fetch_pod_names(container)
        if not pod_regex:
            report.append({
                "Date": report_date,
                "Container Name": container,
                "Max Memory": "N/A",
                "Avg Memory": "N/A",
                "Max CPU": "N/A",
                "Avg CPU": "N/A",
                "Data Points": 0
            })
            continue

        cpu_values, mem_values = fetch_metrics(container, pod_regex, start_time, end_time)
        
        if not cpu_values or not mem_values:
            report.append({
                "Date": report_date,
                "Container Name": container,
                "Max Memory": "Error",
                "Avg Memory": "",
                "Max CPU": "",
                "Avg CPU": "",
                "Data Points": 0
            })
            continue

        stats = {
            "Date": report_date,
            "Container Name": container,
            "Max Memory": max(mem_values),
            "Avg Memory": sum(mem_values)/len(mem_values),
            "Max CPU": max(cpu_values),
            "Avg CPU": sum(cpu_values)/len(cpu_values),
            "Data Points": len(cpu_values)
        }
        report.append(stats)

    return report

def save_reports(report_data):
    """Save report to both CSV and Excel with requested naming format"""
    try:
        df = pd.DataFrame(report_data)
        
        # Reorder columns
        column_order = ["Date", "Container Name", "Max Memory", "Avg Memory", 
                      "Max CPU", "Avg CPU", "Data Points"]
        df = df[column_order]
        
        # Generate filename with current date
        file_date = datetime.now().strftime("%d-%m-%y")
        base_filename = f"{file_date}_newgen_infra"
        
        # Ensure output directory exists
        ensure_output_directory()
        
        # Format percentage columns for CSV
        percentage_cols = ['Max Memory', 'Avg Memory', 'Max CPU', 'Avg CPU']
        for col in percentage_cols:
            df[col] = df[col].apply(format_percentage)
        
        # Save to CSV
        csv_path = os.path.join(OUTPUT_PATH, f"{base_filename}.csv")
        df.to_csv(csv_path, index=False)
        
        # Save to Excel with openpyxl
        excel_path = os.path.join(OUTPUT_PATH, f"{base_filename}.xlsx")
        
        # Create new workbook
        from openpyxl import Workbook
        wb = Workbook()
        
        # Remove default sheet if it exists
        if len(wb.sheetnames) > 0:
            wb.remove(wb.active)
        
        # Create new sheet
        ws = wb.create_sheet('Infra Metrics')
        
        # Write headers
        headers = df.columns.tolist()
        ws.append(headers)
        
        # Write data rows - convert percentages back to decimals for Excel
        for _, row in df.iterrows():
            excel_row = []
            for val, col in zip(row, df.columns):
                if col in percentage_cols and isinstance(val, str) and '%' in val:
                    excel_row.append(float(val.strip('%'))/100)  # Convert to decimal
                else:
                    excel_row.append(val)
            ws.append(excel_row)
        
        # Format percentage columns (0.00% shows as XX.XX% in Excel)
        percentage_format = '0.00%'
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in percentage_cols:
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter][1:]:  # Skip header
                    if cell.value is not None:
                        cell.number_format = percentage_format
        
        # Set column widths
        column_widths = {
            'A': 12,  # Date
            'B': 20,  # Container Name
            'C': 12,  # Max Memory
            'D': 12,  # Avg Memory
            'E': 12,  # Max CPU
            'F': 12,  # Avg CPU
            'G': 12   # Data Points
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Format header row
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Save the workbook
        wb.save(excel_path)
        
        print(f"\nReports saved to {OUTPUT_PATH}:")
        print(f"- CSV file: {os.path.basename(csv_path)}")
        print(f"- Excel file: {os.path.basename(excel_path)}")
        
        return df

    except Exception as e:
        print(f"\nError saving reports: {str(e)}")
        print("Please check if the output directory exists and is writable")
        return None

if __name__ == "__main__":
    print("Generating NewGen Infrastructure Report...")
    report_data = generate_report()
    df = save_reports(report_data)
    
    if df is not None:
        # Format the numeric columns with two decimal places and % sign for display
        percentage_cols = ['Max Memory', 'Avg Memory', 'Max CPU', 'Avg CPU']
        for col in percentage_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f"{float(x.strip('%')):.2f}%" if isinstance(x, str) and '%' in x else x)
        
        print("\nFinal Report:")
        print(df.to_string(index=False))
